"""
Candidate Intake Parser - Extract structured data from candidate application forms.

This module provides deterministic parsing of candidate-provided Excel application
forms (Application Info.xlsx template) into normalized profile dictionaries.

Key features:
- Resilient to blank rows and section headers
- Handles trailing spaces and inconsistent formatting
- Extracts skills with years of experience from "Additional Questions" sheet
- Pure function with no side effects or network calls
"""

import re
from pathlib import Path
from typing import Any, Dict, Optional, Union

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel parsing. Install with: pip install openpyxl"
    )


def parse_application_info_xlsx(path: Union[str, Path]) -> Dict[str, Any]:
    """
    Parse candidate application Excel file into normalized profile.

    Extracts personal information, contact details, work authorization,
    education, and skills with years of experience.

    Args:
        path: Path to Application Info.xlsx file

    Returns:
        Dictionary with keys:
            - first_name (str)
            - last_name (str)
            - email (str)
            - phone (str)
            - address (str)
            - country (str)
            - work_authorization (str)
            - education_level (str)
            - skills_years (dict[str, float|int]): skill name → years

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If required sheets or fields are missing

    Example:
        >>> profile = parse_application_info_xlsx("application.xlsx")
        >>> profile["first_name"]
        'John'
        >>> profile["skills_years"]["Python"]
        5
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Application file not found: {path}")

    workbook = load_workbook(path, data_only=True)

    # Parse main application info
    profile = _parse_main_sheet(workbook)

    # Parse additional questions for skills
    profile["skills_years"] = _parse_skills_sheet(workbook)

    workbook.close()

    return profile


def _parse_main_sheet(workbook) -> Dict[str, Any]:
    """
    Parse the main application sheet for personal and contact information.

    Expects key-value pairs in columns A (label) and B (value).
    Resilient to section headers and blank rows.
    """
    # Try common sheet names
    sheet_names = ["Application Info", "Personal Info", "Sheet1"]
    sheet = None
    for name in sheet_names:
        if name in workbook.sheetnames:
            sheet = workbook[name]
            break

    if sheet is None:
        # Just use first sheet
        sheet = workbook.worksheets[0]

    profile = {
        "first_name": "",
        "last_name": "",
        "email": "",
        "phone": "",
        "address": "",
        "country": "",
        "work_authorization": "",
        "education_level": ""
    }

    # Field name mappings (handle variations)
    field_mappings = {
        "first name": "first_name",
        "firstname": "first_name",
        "last name": "last_name",
        "lastname": "last_name",
        "email": "email",
        "email address": "email",
        "phone": "phone",
        "phone number": "phone",
        "mobile": "phone",
        "address": "address",
        "street address": "address",
        "country": "country",
        "citizenship": "work_authorization",
        "work authorization": "work_authorization",
        "visa status": "work_authorization",
        "education": "education_level",
        "education level": "education_level",
        "highest education": "education_level",
        "degree": "education_level"
    }

    # Parse rows
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 2:
            continue

        label = row[0]
        value = row[1]

        # Skip empty rows
        if not label or not value:
            continue

        # Skip section headers (ALL CAPS with no value or common headers)
        label_str = str(label).strip()
        if label_str.isupper() and len(label_str) > 3:
            # Likely a section header like "PERSONAL INFORMATION"
            continue

        # Normalize label for matching
        normalized_label = label_str.lower().strip()

        # Match to profile field
        if normalized_label in field_mappings:
            field_name = field_mappings[normalized_label]
            profile[field_name] = str(value).strip()

    return profile


def _parse_skills_sheet(workbook) -> Dict[str, Union[int, float]]:
    """
    Parse the Additional Questions sheet for skills and years of experience.

    Looks for questions containing "years" and extracts numeric values.
    Question format: "How many years of [skill] experience do you have?"
    """
    # Try to find Additional Questions sheet
    sheet_names = ["Additional Questions", "Questions", "Skills"]
    sheet = None
    for name in sheet_names:
        if name in workbook.sheetnames:
            sheet = workbook[name]
            break

    if sheet is None:
        # No additional questions sheet found
        return {}

    skills_years = {}

    # Parse rows looking for skill questions
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 2:
            continue

        question = row[0]
        answer = row[1]

        if not question or not answer:
            continue

        # Look for "years of [skill]" pattern
        question_str = str(question).lower()
        if "years" in question_str and "experience" in question_str:
            # Extract skill name
            skill = _extract_skill_name(str(question))
            if skill:
                # Extract numeric years
                years = _extract_years(answer)
                if years is not None:
                    skills_years[skill] = years

    return skills_years


def _extract_skill_name(question: str) -> Optional[str]:
    """
    Extract skill name from question text.

    Examples:
        "How many years of Python experience do you have?" → "Python"
        "Years of experience with Java?" → "Java"
    """
    # Pattern: "years of [skill]" or "years of experience with [skill]"
    patterns = [
        r"years\s+of\s+(?:experience\s+(?:with|in)\s+)?([A-Za-z0-9+#\s]+?)(?:\s+experience|\s+do\s+you|\?|$)",
        r"(?:with|in)\s+([A-Za-z0-9+#\s]+?)\s+(?:experience|\?|$)"
    ]

    question_lower = question.lower()

    for pattern in patterns:
        match = re.search(pattern, question_lower, re.IGNORECASE)
        if match:
            skill = match.group(1).strip()
            # Clean up common artifacts
            skill = skill.replace("?", "").replace("experience", "").strip()
            if skill and len(skill) > 1:
                # Capitalize properly
                return skill.title()

    return None


def _extract_years(answer: Any) -> Optional[Union[int, float]]:
    """
    Extract numeric years from answer.

    Handles:
        - Direct numbers: 5, 3.5
        - Text with numbers: "5 years", "3.5 years"
        - Ranges: "3-5 years" → take midpoint
    """
    if answer is None:
        return None

    # If already a number
    if isinstance(answer, (int, float)):
        return float(answer) if isinstance(answer, float) else int(answer)

    # Extract from string
    answer_str = str(answer).strip().lower()

    # Remove common text
    answer_str = answer_str.replace("years", "").replace("year", "").strip()

    # Try to extract number
    # Handle ranges like "3-5"
    range_match = re.search(r"(\d+\.?\d*)\s*-\s*(\d+\.?\d*)", answer_str)
    if range_match:
        start = float(range_match.group(1))
        end = float(range_match.group(2))
        return (start + end) / 2

    # Extract single number
    number_match = re.search(r"(\d+\.?\d*)", answer_str)
    if number_match:
        num_str = number_match.group(1)
        if "." in num_str:
            return float(num_str)
        else:
            return int(num_str)

    return None
